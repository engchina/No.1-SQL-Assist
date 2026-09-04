"""Tests for SelectAI execution report utilities."""

import json
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from utils.selectai_report_util import (
    ELAPSED_SECONDS_COLUMNS,
    REPORT_COLUMNS,
    append_execution_report,
    create_execution_report_excel,
    disabled_report_download_button,
    execution_report_dataframe,
    format_elapsed_seconds,
    generate_execution_report_download,
    load_execution_report_records,
)


class SelectAiReportUtilTest(unittest.TestCase):
    def test_format_elapsed_seconds_uses_three_decimal_places(self):
        self.assertEqual(format_elapsed_seconds(None), "")
        self.assertEqual(format_elapsed_seconds(0), "0.000")
        self.assertEqual(format_elapsed_seconds(62), "0.062")
        self.assertEqual(format_elapsed_seconds(1500), "1.500")
        self.assertEqual(format_elapsed_seconds(65000), "65.000")

    def test_append_and_load_jsonl_keeps_non_ascii_and_column_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            history_path = Path(temp_dir) / "reports" / "history.jsonl"

            append_execution_report(
                {
                    "画面名": "ユーザー機能: 基本機能",
                    "実行ID": "run-1",
                    "自然言語の質問": "大阪の顧客数を教えて",
                    "カテゴリ予測": "営業",
                    "生成されたSQL": "SELECT COUNT(*) FROM CUSTOMERS",
                    "試行回数": 2,
                    "Select AI経過時間（秒）": "8.700",
                    "SELECT経過時間（秒）": "0.062",
                    "全体経過時間（秒）": "8.800",
                },
                history_path=history_path,
            )
            history_path.write_text(
                history_path.read_text(encoding="utf-8")
                + "{malformed json}\n",
                encoding="utf-8",
            )

            records = load_execution_report_records(history_path)
            self.assertEqual(len(records), 1)
            self.assertEqual(list(records[0].keys()), REPORT_COLUMNS)
            self.assertEqual(records[0]["自然言語の質問"], "大阪の顧客数を教えて")
            self.assertEqual(records[0]["試行回数"], "2")
            self.assertEqual(records[0]["Select AI経過時間（秒）"], "8.700")
            self.assertEqual(records[0]["SELECT経過時間（秒）"], "0.062")

            raw_line = history_path.read_text(encoding="utf-8").splitlines()[0]
            self.assertIn("大阪", raw_line)
            json.loads(raw_line)

    def test_execution_report_excel_uses_stable_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            history_path = temp_root / "history.jsonl"
            output_path = temp_root / "selectai_report.xlsx"
            append_execution_report(
                {
                    "画面名": "開発者機能: チャット・分析",
                    "実行ID": "run-2",
                    "ステータス": "✅ 取得完了",
                    "結果件数": 3,
                    "Select AI経過時間（秒）": "8.700",
                    "SELECT経過時間（秒）": "0.062",
                    "全体経過時間（秒）": "8.800",
                },
                history_path=history_path,
            )

            df = execution_report_dataframe(history_path=history_path)
            self.assertEqual(df.columns.tolist(), REPORT_COLUMNS)
            self.assertEqual(df.iloc[0]["結果件数"], "3")

            created_path = create_execution_report_excel(
                output_path=output_path,
                history_path=history_path,
            )
            self.assertEqual(created_path, output_path)
            self.assertTrue(output_path.exists())

            excel_df = pd.read_excel(output_path)
            self.assertEqual(excel_df.columns.tolist(), REPORT_COLUMNS)
            self.assertEqual(
                excel_df.iloc[0]["画面名"],
                "開発者機能: チャット・分析",
            )
            self.assertEqual(excel_df.iloc[0]["SELECT経過時間（秒）"], 0.062)

            workbook = load_workbook(output_path)
            worksheet = workbook["SelectAI Report"]
            headers = [cell.value for cell in worksheet[1]]
            elapsed_column_index = headers.index("SELECT経過時間（秒）") + 1
            elapsed_cell = worksheet.cell(row=2, column=elapsed_column_index)
            self.assertEqual(elapsed_cell.value, 0.062)
            self.assertEqual(elapsed_cell.number_format, "0.000")

    def test_legacy_elapsed_values_are_normalized_to_seconds(self):
        df = execution_report_dataframe(
            records=[
                {
                    "画面名": "開発者機能: チャット・分析",
                    "Select AI経過時間": "8.7秒",
                    "SELECT経過時間": "62ms",
                    "全体経過時間": "00:09",
                }
            ]
        )
        self.assertEqual(df.columns.tolist(), REPORT_COLUMNS)
        self.assertNotIn("Select AI経過時間", df.columns)
        self.assertEqual(df.iloc[0]["Select AI経過時間（秒）"], "8.700")
        self.assertEqual(df.iloc[0]["SELECT経過時間（秒）"], "0.062")
        self.assertEqual(df.iloc[0]["全体経過時間（秒）"], "9.000")

    def test_execution_report_excel_filters_by_screen(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            history_path = temp_root / "history.jsonl"
            user_output_path = temp_root / "user_report.xlsx"
            missing_output_path = temp_root / "missing_report.xlsx"

            append_execution_report(
                {
                    "画面名": "開発者機能: チャット・分析",
                    "実行ID": "dev-run",
                    "自然言語の質問": "東京の顧客数を教えて",
                },
                history_path=history_path,
            )
            append_execution_report(
                {
                    "画面名": "ユーザー機能: 基本機能",
                    "実行ID": "user-run",
                    "自然言語の質問": "大阪の顧客数を教えて",
                },
                history_path=history_path,
            )

            dev_df = execution_report_dataframe(
                history_path=history_path,
                screen_name="開発者機能: チャット・分析",
            )
            self.assertEqual(len(dev_df), 1)
            self.assertEqual(dev_df.iloc[0]["実行ID"], "dev-run")

            created_path = create_execution_report_excel(
                output_path=user_output_path,
                history_path=history_path,
                screen_name="ユーザー機能: 基本機能",
            )
            self.assertEqual(created_path, user_output_path)
            user_excel_df = pd.read_excel(user_output_path)
            self.assertEqual(len(user_excel_df), 1)
            self.assertEqual(user_excel_df.iloc[0]["実行ID"], "user-run")
            for column in ELAPSED_SECONDS_COLUMNS:
                self.assertIn(column, user_excel_df.columns)

            self.assertIsNone(
                create_execution_report_excel(
                    output_path=missing_output_path,
                    history_path=history_path,
                    screen_name="存在しない画面",
                )
            )
            self.assertFalse(missing_output_path.exists())

    def test_download_button_is_visible_but_disabled_until_report_generated(self):
        initial_button = disabled_report_download_button()
        self.assertTrue(initial_button.visible)
        self.assertFalse(initial_button.interactive)
        self.assertIsNone(initial_button.value)

        with tempfile.TemporaryDirectory() as temp_dir:
            history_path = Path(temp_dir) / "history.jsonl"
            _status, missing_button = generate_execution_report_download(
                history_path=history_path,
                screen_name="ユーザー機能: 基本機能",
            )
            self.assertTrue(missing_button.visible)
            self.assertFalse(missing_button.interactive)
            self.assertIsNone(missing_button.value)

            append_execution_report(
                {
                    "画面名": "ユーザー機能: 基本機能",
                    "実行ID": "user-run",
                    "全体経過時間（秒）": "1.234",
                },
                history_path=history_path,
            )
            _status, generated_button = generate_execution_report_download(
                history_path=history_path,
                screen_name="ユーザー機能: 基本機能",
            )
            self.assertTrue(generated_button.visible)
            self.assertTrue(generated_button.interactive)
            self.assertIsInstance(generated_button.value, dict)
            self.assertTrue(Path(generated_button.value["path"]).exists())


if __name__ == "__main__":
    unittest.main()
